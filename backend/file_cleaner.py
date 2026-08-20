import json
import pprint
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import os

root = tk.Tk()
root.title('Excel combiner')
root.geometry('400x200')

file1 = None
file2 = None
pc_path = r'C:\Users\rjayd\OneDrive\Documents\Excel files\cleaned_file.xlsx' 
laptop_path = r'C:\Users\rjayd\OneDrive\Work\excel_files\cleaned_file.xlsx'
# pc and laptop have different paths, so check first if its the pc_path, if not, use laptop_path
output_path = pc_path if os.path.exists(os.path.dirname(pc_path)) else laptop_path

def select_file1():
    global file1

    file1 = filedialog.askopenfilename(
        title="Select First Excel File",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if file1:
        file1_label.config(text=file1)

def select_file2():
    global file2

    file2 = filedialog.askopenfilename(
        title="Select Second Excel File",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if file2:
        file2_label.config(text=file2)

def combine_files(file1, file2, output):
    if file1 is None and file2 is None:
        print('Waiting for file selection...')
        return

    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    cleaned_file = load_workbook(output_path)

    ws1 = wb1.active
    ws2 = wb2.active
    ws3 = cleaned_file.active

    # helper function to get the data from both files and store them in respective dict's
    def retrive_excel_data(ws):
        data_dict = {}
        for row in ws.iter_rows(min_row=5, max_row=22, min_col=6, max_col=51):
            # getting current row # to use in skipping rows with formula's (19 & 21)
            current_row = row[0].row
            # if current row is 19 or 21, skip and go to next row
            if current_row == 19 or current_row == 21:  
                continue

            print(current_row)
            for cell in row:
                if cell.value is not None:
                    value = cell.value
                    row_num = cell.row
                    col_num = cell.column
                    data_dict[(row_num, col_num)] = value
        return data_dict

    data_dict1 = retrive_excel_data(ws1)
    data_dict2 = retrive_excel_data(ws2)

    combined_data = {}

    # iterate thru both dict's, add values with matching row/col key's, then append to dict
    for key, value in data_dict1.items():
        for key2, value2 in data_dict2.items():
            if key2 == key:
                combined_data[(key2)] = value + value2
                break

    # print(combined_data)

    # iterate over dict and add into new excel file
    for (row, col), value in combined_data.items():
        
        ws3.cell(
            row=row,
            column=col,
            value=value
        )

    print('File created successfully!')
    
    cleaned_file.save(output_path)

# file 1
tk.Button(
    root,
    text='Select File 1',
    command=select_file1
).pack(pady=10)

file1_label = tk.Label(root, text="No file selected")
file1_label.pack()

tk.Button(
    root,
    text='Select File 2',
    command=select_file2
).pack(pady=10)

file2_label = tk.Label(root, text="No file selected")
file2_label.pack()

# lambda runs combine_files once button is clicked
tk.Button(
    root,
    text='Combine Files',
    command= lambda: combine_files(file1, file2, output_path)
).pack(pady=15)

# start tkinter
root.mainloop()